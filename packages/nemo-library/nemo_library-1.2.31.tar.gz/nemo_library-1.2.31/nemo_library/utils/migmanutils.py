import importlib
import logging
import os
import tempfile
import openpyxl
import pandas as pd

from nemo_library.features.config import Config
from nemo_library.features.fileingestion import ReUploadFile


def initializeFolderStructure(
    project_path: str,
) -> None:

    folders = ["templates", "mappings", "srcdata", "other"]
    for folder in folders:
        os.makedirs(os.path.join(project_path, folder), exist_ok=True)


def getMappingFilePath(projectname: str, local_project_path: str) -> str:
    return os.path.join(local_project_path, "mappings", f"{projectname}.csv")


def load_database() -> pd.DataFrame:
    with importlib.resources.open_binary(
        "nemo_library.templates", "migmantemplates.pkl"
    ) as file:
        df = pd.read_pickle(file)

    return df


def getProjectName(project: str, addon: str, postfix: str) -> str:
    return f"{project}{" " + addon if addon else ""}{(" (" + postfix + ")") if postfix else ""}"


def getNEMOStepsFrompAMigrationStatusFile(file: str) -> list[str]:
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook["Status Datenübernahme"]

    data = []
    for row in worksheet.iter_rows(
        min_row=10, max_row=300, min_col=1, max_col=10, values_only=True
    ):
        data.append(row)

    # Create a DataFrame from the extracted data
    columns = [
        worksheet.cell(row=9, column=i).value for i in range(1, 11)
    ]  # Headers in row 9
    dataframe = pd.DataFrame(data, columns=columns)

    # Drop rows where "Importreihenfolge" is NaN or empty
    if "Importreihenfolge" in dataframe.columns:
        dataframe = dataframe.dropna(subset=["Importreihenfolge"])
    else:
        raise ValueError("The column 'Importreihenfolge' does not exist in the data.")

    if "Name des Importprograms / Name der Erfassungsmaske" in dataframe.columns:
        nemosteps = dataframe[dataframe["Migrationsart"] == "NEMO"][
            "Name des Importprograms / Name der Erfassungsmaske"
        ].to_list()

        nemosteps = [x.title().strip() for x in nemosteps]
        replacements = {
            "European Article Numbers": "Global Trade Item Numbers",
            "Part-Storage Areas Relationship": "Part-Storage Areas Relationships",
            "Sales Tax Id": "Sales Tax ID",
            "Mrp Parameters": "MRP Parameters",
            "Sales Units Of Measure": "Sales Units of Measure",
            "Standard Boms (Header Data)": "Standard BOMs (Header Data)",
            "Standard Boms (Line Data)": "Standard BOMs (Line Data)",
            "Routings (Standard Boms)": "Routings (Standard BOMs)",
            "Bills Of Materials For Operations (Routings Production)": "Bills of Materials for Operations (Routings Production)"
        }

        nemosteps = [
            replacements[item] if item in replacements else item for item in nemosteps
        ]

        return nemosteps
    else:
        raise ValueError(
            "The column 'Name des Importprograms / Name der Erfassungsmaske' does not exist in the data."
        )


def upload_dataframe(config: Config, project_name: str, df: pd.DataFrame):

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_file_path = os.path.join(temp_dir, "tempfile.csv")

        df.to_csv(
            temp_file_path,
            index=False,
            sep=";",
            na_rep="",
            encoding="UTF-8",
        )
        logging.info(
            f"dummy file {temp_file_path} written for project '{project_name}'. Uploading data to NEMO now..."
        )

        ReUploadFile(
            config=config,
            projectname=project_name,
            filename=temp_file_path,
            update_project_settings=False,
            version=3,
            datasource_ids=[{"key": "datasource_id", "value": project_name}],
        )
        logging.info(f"upload to project {project_name} completed")
