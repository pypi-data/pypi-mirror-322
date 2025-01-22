import base64
import json
import os
import re
import time
import getpass
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO, StringIO, IOBase

import numpy as np
import pandas as pd
import openpyxl as xl
import requests
from PIL import Image
from requests_toolbelt import MultipartEncoder
import logging

from ..settings import HUE_DOWNLOAD_BASE_URL, EXCEL_ENGINE
from ..decorators import retry, ensure_login
from .. import logger


class HueDownload(requests.Session):

    def __init__(self,
                 username: str = None,
                 password: str = None,
                 verbose: bool = False):
        self.base_url = HUE_DOWNLOAD_BASE_URL

        self.username = username
        self._password = password
        self.verbose = verbose
        self.log = logging.getLogger(__name__ + ".HueDownload")
        if verbose:
            logger.set_stream_log_level(self.log, verbose)

        self.log.debug("loading img_dict")
        self.benchmark_imgs = np.load(os.path.join(os.path.dirname(__file__), "img_dict.npy"), allow_pickle=True).item()
        super(HueDownload, self).__init__()

        self.login(self.username, self._password)

    @retry(__name__)
    def _login(self, username, password):
        self.headers.update({
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json",
            "Referer": "http://10.19.185.103:8015/login?redirect=%2Fdashboard",
            "User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64) "
                          "AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/76.0.3809.100 Safari/537.36"
        })
        login_url = self.base_url + "/auth/login"
        form_data = dict(username=username,
                         password=password,
                         code=self.code,
                         uuid=self.uuid)

        res = self.post(login_url,
                        data=json.dumps(form_data))
        return res

    def login(self, username=None, password=None):
        self.is_logged_in = False
        self.username = username or self.username
        self._password = password or self._password
        if self.username is None and self._password is None:
            raise ValueError("please provide username and password")

        if self.username is None and self._password is not None:
            raise KeyError("username must be specified with password")

        if self.username is not None and self._password is None:
            self._password = getpass.getpass("Please provide HueDownload password: ")

        self.log.debug(f"logging in for user [{self.username}]")
        self.id_answer()
        res = self._login(self.username, self._password)
        r_json = res.json()
        if "status" in r_json.keys():
            if r_json["status"] == 400 and r_json["message"] == "验证码错误":
                raise ConnectionError("captcha guess failed")

            self.log.error(res.text)
            raise RuntimeError(r_json["message"])

        self.log.info('login succeeful [%s] at %s'
                      % (self.username, self.base_url))
        self.is_logged_in = True
        self.headers["Authorization"] = "Bearer " + r_json["token"]

    def get_column(self, table_name):
        res = self._get_column(table_name=table_name)
        columns = [desc["name"] for desc in res.json()]
        return columns

    def download_data(self, table_name, reason, col_info=' ', limit=None, columns=None, Decode_col=[]):
        """
        table_name 必填，需要下载的表名
        reason 必填，下载事由
        col_info 选填,默认值' ',
        limit 选填，默认值None，下载条数不填则全部下载，最多10万行
        columns 选填，默认值None，不填则全部下载
        Decode_col 选填，默认值[]， 不填则不解密
        """
        self.log.warning("download_data is depreciated and won't be maintained in the future,"
                         "please instead use 'download'")
        download_info = {}
        self.headers['Referer'] = 'http://10.19.185.103:8015/ud/downloadInfo'
        self.headers['Content-Type'] = 'application/json'

        if 'Authorization' not in self.headers.keys():
            self.login(self.username, self._password)
        url = self.base_url + '/api/downloadInfo'

        if columns is None:
            columns = self.get_column(table_name)
            if len(columns) > 200:
                th = ThreadPoolExecutor(max_workers=3)
                results = []

                for i in range(0, int(len(columns) / 200) + 1):
                    start_num = i * 200
                    end_num = (i + 1) * 200 - 1
                    if end_num > len(columns) - 1:
                        end_num = len(columns)

                    temp_column = columns[start_num:end_num]
                    temp_reason = reason + ' part ' + str(i)
                    # print(temp_column)
                    results.append(th.submit(self.download_data, table_name, temp_reason, columns=temp_column))
                    # print('sub')
                cnt = 1
                # print('submited')
                for result in results:
                    temp_df = result.result()
                    if cnt == 1:
                        result_df = temp_df
                        cnt += 1

                    else:
                        result_df = pd.merge(result_df, temp_df, left_index=True, right_index=True)

                return result_df

        if limit is not None:
            download_info['downloadLimit'] = limit

        download_info['downloadTable'] = table_name
        download_info['downloadColumns'] = columns
        download_info['reason'] = reason

        if Decode_col is not None:
            download_info['downloadDecryptionColumns'] = Decode_col
        download_info['columnsInfo'] = col_info

        r = requests.post(url, data=json.dumps(download_info))
        r = r.json()
        # print(r)
        if r['status'] != 0:
            print(r['message'])
            return
        t_sec = 30
        t_try = 100
        t_tol = t_sec * t_try
        job_id = r['id']
        tag = 0
        for i in range(t_try):
            print('waiting %3d/%d...' %
                  (t_sec * i, t_tol) + '\r', end='')
            r = requests.get(self.base_url + '/api/downloadInfo?page=0&size=10&sort=id,desc')
            task_list = r.json()['content']
            for task in task_list:
                if task['id'] == job_id and task['status'] == 3:
                    tag = 1
                    break
            if tag == 1:
                break
            time.sleep(t_sec)
        if col_info == ' ':
            csv_header = 0
        else:
            csv_header = 1
        r = requests.get(self.base_url + '/api/downloadInfo/downloadData?id=' + str(job_id))
        r = pd.read_csv(StringIO(r.text), header=csv_header)
        return r

    def download(self,
                 table: str,
                 reason: str,
                 columns: list = None,
                 column_names: list = None,
                 decrypt_columns: list = None,
                 limit: int = None,
                 path: str = None,
                 wait_sec: int = 5,
                 timeout: float = float("inf"),
                 **info_kwargs
                 ):
        """
        a refactored version of download_data from WxCustom
        specify table information and load or download to local

        :param table: table name on Hue (database name is required)
        :param reason:  reason of downloading
        :param columns: specify which of the columns in table to download from Hue,
                        default to all columns
        :param column_names: rename column names if needed
        :param decrypt_columns: columns to be decrypted
        :param limit: the maximum number of records to be downloaded
                      default to all records
        :param path: output csv file if specified.
                     default to return Pandas.DataFrame
                     this is designed to download large table without using up memory
        :param wait_sec: time interval while waiting server for preparing for download
                         default to 5 seconds
        :param timeout: maximum seconds to wait for the server preparation
                       default to wait indefinitely
        :param info_kwargs: to modify get_info_by_id parameters, add argument pairs here
                            (useful when downloadables cannot be found in just one page)

        :return: Pandas.DataFrame if path is not specified,
                 otherwise output a csv file to path and return None
        """
        if columns is None:
            columns = self.get_column(table)
        if decrypt_columns is not None:
            columns = pd.unique(columns + decrypt_columns).tolist()
        if column_names:
            if len(columns) != len(column_names):
                self.log.warning(f"length of table columns({len(columns)}) "
                                 f"mismatch with column_names({len(column_names)}), rename skipped")
                column_names = None

        res = self._download(
            table=table,
            reason=reason,
            columns=columns,
            decrypt_columns=decrypt_columns,
            limit=limit)
        r_json = res.json()

        error_msg = f"cannot download {table}, please check table name and (decrypt) columns"
        if r_json["status"] != 0:
            self.log.error(error_msg)
            raise RuntimeError(error_msg)

        download_id = r_json["id"]
        start_time = time.perf_counter()
        while time.perf_counter() - start_time < timeout:
            time.sleep(wait_sec)
            download_info = self.get_info_by_id(download_id, info_type="download", **info_kwargs)
            if download_info["status"] == 0:
                # status: submit
                self.log.info(f"prepare {table} elapsed: {time.perf_counter() - start_time:.2f}/{timeout} secs")
                continue
            if self.verbose:
                print()
            if download_info["status"] == 1:
                # status: failed
                raise RuntimeError(error_msg)
            if download_info["status"] == 3:
                # status: success
                return self.download_by_id(download_id=download_id, path=path, column_names=column_names)
            else:
                self.log.error(f"can't resolve download info: {download_info}")
                raise RuntimeError(f"can't resolve download info: {download_info}")

        if self.verbose:
            print()
        self.log.error(f"download {table} timed out")
        return TimeoutError(f"download {table} timed out")

    @ensure_login
    def upload_data(self,
                    file_path: str,
                    reason: str,
                    column_names: str = '1',
                    encrypt_columns: str = '',
                    wait_sec: int = 5,
                    timeout: float = float("inf")
                    ):
        """
            file_path  必填，需要上传文件位置
            reason 必填，上传事由
            uploadColumnsInfo 选填，默认写1，可用作备注，与上传数据无关
            uploadEncryptColumns 选填，默认'',需要加密的列，多个用逗号隔开
        """
        self.log.warning("upload_data is depreciated and won't be maintained in the future,"
                         "please instead use 'upload'")
        if re.findall('\.csv$|\.xlsx?$', file_path):
            # instead of read all data in memory using pd.read_...
            # read only necessary column info and row count
            wb = xl.load_workbook(file_path, read_only=True)
            sheet = wb.worksheets[0]

            columns = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
            if len(columns) == 0:
                raise RuntimeError(f'get empty list of column name from input file {file_path}')

            rows = sheet.max_row
            if rows == 0:
                raise RuntimeError(f'get empty rows of input file {file_path}')
        else:
            raise RuntimeError('data format is not supported yet! please upload csv or xlsx with english title')

        wb.close()
        buffer = open(file_path, "rb")
        res = self._upload(file_buffer=buffer,
                           reason=reason,
                           columns=columns,
                           column_names=column_names,
                           encrypt_columns=encrypt_columns,
                           nrows=rows)
        buffer.close()
        id_ = res.json()['id']

        error_msg = f"cannot upload {file_path}, please check table name and (encrypt) columns"
        start_time = time.perf_counter()
        while time.perf_counter() - start_time < timeout:
            time.sleep(wait_sec)
            upload_info = self.get_info_by_id(id_, info_type="upload")
            if upload_info["status"] == 0:
                # status: submit
                self.log.info(f"setup upload elapsed: {time.perf_counter() - start_time:.2f}/{timeout} secs")
                continue
            if upload_info["status"] == 1:
                # status: failed
                self.log.error(f"RuntimeError: cannot upload {file_path}")
                raise RuntimeError(error_msg)
            if upload_info['status'] == 3:
                return upload_info['rsTable']

        self.log.error(f"upload {file_path} timed out")
        return TimeoutError(f"upload {file_path} timed out")

    def upload(self,
               data,
               reason: str,
               columns: list = None,
               column_names: list = None,
               encrypt_columns: list = None,
               nrows: int = None,
               wait_sec: int = 5,
               timeout: float = float("inf"),
               **info_kwargs
               ):
        """
        a refactored version of upload_data from WxCustom
        parse upload data and call upload API, if success, return uploaded table name.

        :param data: pandas.DataFrame, pandas.Series or path str to xlsx, xls or csv file
        :param reason: str, upload reason
        :param columns: list, list of columns to upload
        :param column_names: list, list of column with respective to their alias,
                            must be as same length as columns
        :param encrypt_columns: list, list of columns to encrypt during upload
        :param nrows: number of rows to upload, default to be -1, all rows
        :param wait_sec: time interval while waiting server for preparing for upload
                         default to 5 seconds
        :param timeout: maximum seconds to wait for the server preparation
                        default to wait indefinitely
        :param info_kwargs: to modify get_info_by_id parameters, add argument pairs here
                            (useful when upload contents cannot be found in just one page)

        :return: str, name of uploaded table
        """

        if isinstance(data, (pd.DataFrame, pd.Series)):
            buffer = BytesIO()
            buffer.name = "pd.DataFrame.xlsx"
            data.to_excel(buffer, index=False)
            columns, nrows = columns or data.columns, nrows or data.shape[0]
        elif isinstance(data, str) and re.findall('\.xlsx$|\.xls$|\.xlsm$|\.xltx$|\.xltm$', data):
            # instead of read all data in memory using pd.read_...
            # read only necessary column info and row count
            wb = xl.load_workbook(data, read_only=True)
            sheet = wb.worksheets[0]

            columns = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
            if len(columns) == 0:
                raise RuntimeError(f'get empty list of column name from input file {data}')

            nrows = nrows or sheet.max_row
            if nrows == 0:
                raise RuntimeError(f'get empty rows of input file {data}')

            wb.close()
            buffer = open(data, "rb")
        elif isinstance(data, str) and re.findall('\.csv$', data):
            buffer = open(data, "rb")
            data = pd.read_csv(data)
            columns, nrows = columns or data.columns, nrows or data.shape[0]
        else:
            raise RuntimeError('data format is not supported yet,'
                               ' please upload DataFrame, csv or xlsx with english title')
        if encrypt_columns:
            set_encrypt_columns, set_columns = set(encrypt_columns), set(columns)
            if not set_encrypt_columns.issubset(set_columns):
                err_msg = f"encrypt column" \
                          f" {','.join(set_encrypt_columns.difference(set_columns))}" \
                          f" not in {buffer.name}"
                buffer.close()
                self.log.error(err_msg)
                raise ValueError(err_msg)

        res = self._upload(file_buffer=buffer,
                           reason=reason,
                           columns=columns,
                           column_names=','.join(column_names) if column_names else '',
                           encrypt_columns=','.join(encrypt_columns) if encrypt_columns else '',
                           nrows=nrows or -1)
        buffer.close()
        id_ = res.json()['id']
        error_msg = f"cannot upload {buffer.name}, please check table name and (encrypt) column names"
        start_time = time.perf_counter()
        while time.perf_counter() - start_time < timeout:
            time.sleep(wait_sec)
            upload_info = self.get_info_by_id(id_, info_type="upload", **info_kwargs)
            if upload_info["status"] == 0:
                # status: submit
                self.log.info(f"setup upload elapsed: {time.perf_counter() - start_time:.2f}/{timeout} secs")
                continue
            if upload_info["status"] == 1:
                # status: failed
                self.log.error(f"RuntimeError: cannot upload {buffer.name}")
                raise RuntimeError(error_msg)
            if upload_info['status'] == 3:
                return upload_info['rsTable']

        self.log.error(f"upload {buffer.name} timed out")
        return TimeoutError(f"upload {buffer.name} timed out")

    def get_info_by_id(self, id_: int, info_type, **kwargs):
        if info_type == 'upload' or info_type == 1:
            func_info = self._get_upload_info
        elif info_type == 'download' or info_type == 0:
            func_info = self._get_download_info
        else:
            raise TypeError(f"expecting info_type to be either 'upload' or 1, 'download' or 0, got {info_type}")

        pages = kwargs["page"] if "page" in kwargs else 0
        for p in range(pages + 1):
            kwargs["page"] = p
            res = func_info(**kwargs)
            r_json = res.json()
            for content in r_json["content"]:
                if content["id"] == id_:
                    return content

        raise LookupError(f"cannot get info with download id: {id_}")

    def download_by_id(self, download_id, column_names=None, path=None):
        start_time = time.perf_counter()
        buffer = self._download_by_id(download_id)
        if path is None:
            df = pd.read_csv(StringIO(buffer.text))
            if column_names:
                df.columns = column_names

            return df

        if not isinstance(path, str):
            raise TypeError(f"path should be string, got {type(path)}")

        suffix = path.rpartition(".")[-1]
        if suffix in ("xlsx", "xls", "xlsm"):
            df = pd.read_csv(StringIO(buffer.text))
            if column_names:
                df.columns = column_names

            df.to_excel(path, index=False, engine=EXCEL_ENGINE)
        elif column_names:
            df = pd.read_csv(StringIO(buffer.text))
            df.columns = column_names
            df.to_csv(path, index=False, encoding="utf-8")
        else:
            with open(path, "wb") as f:
                for chunk in buffer.iter_content(chunk_size=8192):
                    f.write(chunk)

        self.log.info(f"download finished in {time.perf_counter() - start_time:.3f} secs")

    @ensure_login
    @retry(__name__)
    def _get_column(self, table_name):
        self.log.debug(f"getting columns for {table_name}")
        url = self.base_url + '/api/hive/getColumns?tableName=' + table_name
        res = self.get(url)
        return res

    @ensure_login
    @retry(__name__)
    def _download(self,
                  table: str,
                  reason: str,
                  columns: list,
                  column_names: str = '',
                  decrypt_columns: list = None,
                  limit: int = ''
                  ):

        self.log.debug(f"downloading {table}")
        url = self.base_url + '/api/downloadInfo'
        self.headers['Content-Type'] = 'application/json'

        res = self.post(url, data=json.dumps({
            "columnsInfo": column_names,
            "downloadColumns": columns,
            "downloadDecryptionColumns": decrypt_columns or [],
            "downloadLimit": str(limit) if limit else '',
            "downloadTable": table,
            "reason": reason
        }))
        return res

    @ensure_login
    @retry(__name__)
    def _upload(self,
                file_buffer: IOBase,
                reason: str,
                columns: list,
                column_names: str,
                encrypt_columns: str,
                nrows: int = -1
                ):

        self.log.info(f"uploading {file_buffer.name}")
        url = self.base_url + '/api/uploadInfo/upload'
        upload_info = {'reason': reason,
                       'uploadColumnsInfo': column_names,
                       'uploadEncryptColumns': encrypt_columns,
                       "uploadColumns": ",".join(columns),
                       'uploadRow': str(nrows)}

        if file_buffer.name == "pd.DataFrame.xlsx":
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            file = (os.path.basename(file_buffer.name), file_buffer, content_type)
        else:
            file = (os.path.basename(file_buffer.name), file_buffer)
        upload_info['file'] = file
        data = MultipartEncoder(fields=upload_info)
        self.headers['Content-Type'] = data.content_type
        res = self.post(url, data=data)
        return res

    @ensure_login
    @retry(__name__)
    def _get_download_info(self,
                           page=0,
                           size=10,
                           sort="id,desc"
                           ):

        self.log.debug(f"getting download info")
        url = self.base_url + '/api/downloadInfo'
        res = self.get(url, params={
            "page": page,
            "size": size,
            "sort": sort
        })
        return res

    @ensure_login
    @retry(__name__)
    def _get_upload_info(self,
                         page=0,
                         size=10,
                         sort="id,desc"
                         ):

        self.log.debug(f"getting upload info")
        url = self.base_url + '/api/uploadInfo'
        res = self.get(url, params={
            "page": page,
            "size": size,
            "sort": sort
        })
        return res

    @retry(__name__)
    def _download_by_id(self, download_id: int):
        self.log.debug(f"downloading by id {download_id}")
        url = self.base_url + '/api/downloadInfo/downloadData'
        res = self.get(
            url,
            params={"id": download_id},
            stream=True
        )
        return res

    def kill_app(self, app_id):
        """
        kill a YARN application

        :param app_id: str or string iterable
        :return: server response context
        """

        if isinstance(app_id, str):
            res = self._kill_app(app_id)
            r_json = res.json()
            if r_json["status"] != 1:
                raise RuntimeError(res.text)
        else:
            # let it fail if app_id is not iterable
            for app in app_id:
                res = self._kill_app(app)
                r_json = res.json()
                if r_json["status"] != 1:
                    raise RuntimeError(res.text)

    @ensure_login
    @retry(__name__)
    def _kill_app(self, app_id: str):
        self.log.debug(f"killing app {app_id}")
        url = self.base_url + '/api/killJobHist'
        res = self.post(url, data=json.dumps({
            "appId": app_id,
            "createTime": "",
            "id": "",
            "ip": "",
            "reason": "",
            "status": "",
            "username": ""
        }))
        self.log.debug(f"_kill_app responds: {res.text}")
        return res

    def base64_pil(self):
        self.img = base64.b64decode(self.img)
        self.img = Image.open(BytesIO(self.img)).convert("L")
        self.img = np.array(self.img)

    def clear_edged(self, img):
        temp = np.sum(img, axis=0)
        # crop image, drop empty vertical pixels
        img = img[:, temp < img.shape[0]]
        return img

    def compare_img(self, imga, imgb):
        # 1 means a complete mismatch, 0 means perfect match
        score = 1.
        ax, ay = imga.shape
        bx, by = imgb.shape

        for i in range(0, abs(ay - by) + 1):
            if ay >= by:
                tmp_score = (imga[:, i:by + i] ^ imgb).sum() / (bx * by)
            else:
                tmp_score = (imga ^ imgb[:, i:ay + i]).sum() / (ax * ay)

            if tmp_score < score:
                score = tmp_score

        return score

    def match_img(self, img):
        score = 1
        result = -1
        for i, benchmark_img in self.benchmark_imgs.items():
            tmp_score = self.compare_img(img, benchmark_img)
            if tmp_score < score:
                score = tmp_score
                result = i
        return result

    @retry(__name__)
    def _get_img(self):
        code_url = self.base_url + "/auth/code"
        res = self.get(code_url)
        return res

    @retry(__name__)
    def id_answer(self):
        code = self._get_img().json()
        self.img = re.sub("data:image/png;base64,", "", code["img"]).replace("%0A", "\n")
        self.uuid = code["uuid"]

        self.base64_pil()
        self.img[self.img <= 180] = 0
        self.img[self.img > 180] = 1
        p1 = self.clear_edged(self.img[:, :24])
        p2 = self.clear_edged(self.img[:, 25:50])
        p3 = self.clear_edged(self.img[:, 51:70])

        num1 = int(self.match_img(p1))
        method = self.match_img(p2)
        num2 = int(self.match_img(p3))

        if method == "+":
            answer = num1 + num2
        elif method == "-":
            answer = num1 - num2
        elif method == "x":
            answer = num1 * num2

        self.code = answer
