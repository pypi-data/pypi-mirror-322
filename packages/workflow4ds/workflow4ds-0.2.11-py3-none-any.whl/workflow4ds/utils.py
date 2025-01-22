import re
import os
import socket
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def human_readable_size(size, decimal_places=2):
    for unit in ['B', 'KiB', 'MiB', 'GiB', 'TiB', 'PiB']:
        if size < 1024. or unit == 'PiB':
            break
        size /= 1024.

    return f"{size:.{decimal_places}f} {unit}"


def get_ip():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.settimeout(0)
    try:
        # doesn't even have to be reachable
        s.connect(('10.254.254.254', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1'
    finally:
        s.close()

    return IP


def reduce_mem_usage(df: pd.DataFrame):
    """
        reduce memory usage by shrinking data types
    """
    for col in df.columns:
        col_type = df[col].dtype
        if re.findall('float|int', str(col_type)):
            c_min = df[col].min()
            c_max = df[col].max()

            if str(col_type)[:3] == 'int':
                if c_min > np.iinfo(np.int8).min and c_max < np.iinfo(np.int8).max:
                    df[col] = df[col].astype(np.int8)
                elif c_min > np.iinfo(np.int16).min and c_max < np.iinfo(np.int16).max:
                    df[col] = df[col].astype(np.int16)
                elif c_min > np.iinfo(np.int32).min and c_max < np.iinfo(np.int32).max:
                    df[col] = df[col].astype(np.int32)
                elif c_min > np.iinfo(np.int64).min and c_max < np.iinfo(np.int64).max:
                    df[col] = df[col].astype(np.int64)
            else:
                if c_min > np.finfo(np.float16).min and c_max < np.finfo(np.float16).max:
                    df[col] = df[col].astype(np.float16)
                elif c_min > np.finfo(np.float32).min and c_max < np.finfo(np.float32).max:
                    df[col] = df[col].astype(np.float32)
                else:
                    df[col] = df[col].astype(np.float64)
        else:
            continue

    return df


def read_file_in_chunks(file_object, block_size, chunks=-1):
    """
    Lazy function (generator) to read a file piece by piece.
    Default chunk size: 25MB.
    """
    i = 1
    while chunks:
        data = file_object.read(block_size)
        if len(data) < block_size:
            yield -1, data
            return

        yield i, data
        i += 1
        chunks -= 1


def append_df_to_csv(filename, df: pd.DataFrame, **to_csv_kwargs):
    """
        Append a DataFrame [df] to existing csv file [filename].
    If [filename] doesn't exist, then this function will create it.

    @param filename: csv path (Example: '/path/to/file.csv')
    @param df: DataFrame to save to file
    @param to_csv_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                          [can be a dictionary]
    @return: None
    """

    if 'header' in to_csv_kwargs:
        to_csv_kwargs.pop('header')
    if 'mode' in to_csv_kwargs:
        to_csv_kwargs.pop('mode')

    if not os.path.isfile(filename):
        df.to_csv(filename, mode='w', header=True, **to_csv_kwargs)
    else:
        df.to_csv(filename, mode='a', header=False, **to_csv_kwargs)


def append_df_to_excel(filename, df: pd.DataFrame,
                       sheet_name='Sheet1',
                       startrow=None,
                       truncate_sheet=False,
                       date_format=None,
                       datetime_format=None,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore parameters if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    if 'header' in to_excel_kwargs:
        to_excel_kwargs.pop('header')

    if pd.__version__ >= '1.4.0':
        book = load_workbook(filename)
        if startrow is None and sheet_name in book.sheetnames:
            startrow = book[sheet_name].max_row
        elif startrow is None:
            startrow = 0

        with pd.ExcelWriter(
            filename,
            engine='openpyxl',
            mode='a',
            date_format=date_format,
            datetime_format=datetime_format,
            if_sheet_exists="replace" if truncate_sheet else "overlay"
            ) as writer:
            df.to_excel(writer,
                        sheet_name=sheet_name,
                        startrow=startrow,
                        header=sheet_name not in book.sheetnames,
                        **to_excel_kwargs)
    else:
        writer = pd.ExcelWriter(filename,
                                engine='openpyxl',
                                mode='a',
                                date_format=date_format,
                                datetime_format=datetime_format,
                                if_sheet_exists="overlay")

        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer,
                    sheet_name,
                    startrow=startrow,
                    header=None,
                    **to_excel_kwargs)

        # save the workbook
        writer.save()
