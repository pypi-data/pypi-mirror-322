from openpyxl import load_workbook

class Excel(object):
    '''
    Excel复杂加工包装
    文件名需要使用绝对路径，例如：c:\\tool\\file_name.xlsx
    sheet可以作为str或者int类型，str需要传入sheet的名字，int需要传入sheet的index

    '''
    def __init__(self, file: str , sheet_index: str|int, ) -> object:
        # load workbook
        self.workbook = load_workbook(file)
        # select workbook
        self.sheet = self.workbook[sheet_index]

    
    def read_data_cols(self, head_index: int, cols: str|list, ) -> list:
        """
        获取表格特定几个列的所有数据
        

        :param head_index: 表头占用的行数
        :type head_index: int

        :param cols: 如果只是一个列, 直接写列的名字字符串, 多个的话, 用list来输入
        :type cols: tr|list

        :rtype: list
        """
        result = []
        if isinstance(cols, str):
            col_names = [cols]
        elif isinstance(cols, list):
            col_names = cols
        else:
            raise TypeError("Unsupported type. Please provide a string or list.")
        column_index = []
        for col in col_names:
            column_index.append({'col': col, 'col_index': self.find_value(col)['col_index']})

        for row in self.sheet.iter_rows(min_row=head_index+1, max_col=max(item['col_index'] for item in column_index), values_only=True):
            obj = {}
            for col_index in column_index:
                obj[col_index['col']] = row[col_index['col_index']-1]
            result.append(obj)
        return result
        

    def find_value(self, search_value: str, ) -> object:
        row_index = 1
        for row in self.sheet.iter_rows(values_only=True):
            col_index = 1
            for cell in row:
                if cell == search_value:
                    return {'row_index':row_index, 'col_index':col_index}
                col_index = col_index + 1
            row_index = row_index + 1
        return {'row_index':None, 'col_index':None}

    def close(self):
        self.workbook.close()