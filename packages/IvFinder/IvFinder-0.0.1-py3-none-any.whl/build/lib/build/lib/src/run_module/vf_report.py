import os
import sys
import time
BASE_DIR=  os.path.dirname(os.path.dirname( os.path.abspath(__file__) ))                   
# 将这个路径添加到环境变量中。
sys.path.append( BASE_DIR  )

import numpy as np
import pandas as pd
import subprocess
import traceback
import argparse
from .find_options import find_advanced_options
import utilities
import click
from configs import options
from configs import config

import openpyxl

from . import classify_radical
from . import virus_divide, draw_krona
import logging
from Bio import SeqIO
from ete3 import NCBITaxa
ncbi = NCBITaxa()
logger=logging.getLogger(__name__)

@click.command(
    #cls=GlobalArgs,
    name = 'report', help='Classification and Generate abundance report',
)
@click.help_option('--help', '-h', help='Show this message and exit')
@click.option(
    '-p', '--processes',
    type=int,
    default=config.processes,
    metavar='<int>',
    help = 'number of processes. [default: '+str(config.processes)+ ']'
)
@options.add_options(options.report_options_list)
@options.add_options(options.CAT_options_list)
@options.add_options(options.global_options_list)


def main(*args, **kwargs):
    start = time.time()
    args = argparse.Namespace()
    for key, value in kwargs.items():
        setattr(args, key, value)

    temp_output_files = []
    utilities.setup_global_settings(args)
    utilities.setup_logging(args)
    utilities.update_configuration(args, 'report', temp_output_files, 'separated')

    generate_report(
        args.virus_anno, args.nonvirus_anno,
        args.orf, args.contigs, args.quant,
        args.cluster_rep, args.cluster_file,
        args.one_minus_r,
        args.report_dir,
        args.report_log,
        args.processes
    )

    for file in temp_output_files:
        utilities.remove_file(file)

    end = time.time()
    print('MVF Generate Report Model Complete in %s s'%(str(end-start)))
    logger.info('MVF Generate Report Model Complete in %s s'%(str(end-start)))

def write_output(taxon_file, abundance_file, contig_file, orf_file, outdir, name, report_log): 
    utilities.create_directory(outdir)
    if taxon_file:
        generate_virus_table(
            taxon_file,
            os.path.join(outdir, name + '_tax.txt'),
            os.path.join(outdir, name + '_tax.xlsx')
        )
        generate_virus_contig(
            contig_file,
            taxon_file,
            os.path.join(outdir, name + '_contig.fasta')
        )
        generate_virus_gene(
            orf_file,
            taxon_file,
            os.path.join(outdir, name + '_gene.fasta')
        )

    if abundance_file:
        generate_virus_table(
            abundance_file,
            os.path.join(outdir, name + '_abundance.txt'),
            os.path.join(outdir, name + '_abundance.xlsx')
        )
        draw_krona.generate_krona(
            abundance_file,
            name,
            outdir,
            logger,
            report_log
        )

def generate_virus_table(taxon_file, output_txt, output_excel):
    with open(output_txt, 'w') as outf:
        keys = list(taxon_file.keys())
        headers = list(taxon_file[keys[0]].keys())

        workbook = openpyxl.Workbook()
        worksheet =  workbook.active
        outf.write('\t'.join(x.title() for x in headers)+'\n')
        for i, header in enumerate(headers):
            worksheet.cell(row = 1, column = i + 1, value = header.title())

        for i, key in enumerate(keys):
            info = []
            value = taxon_file[key]
            #print(value)
            for j, header in enumerate(headers):
                info.append(value[header])
                worksheet.cell(row = i + 2, column = j + 1, value = value[header])
            outf.write('\t'.join(str(x) for x in info)+'\n')
        workbook.save(output_excel)

def generate_virus_contig(contig_file, virus2tax, output_contig):
    contig_parse = SeqIO.parse(contig_file, 'fasta')
    virus_contig = []
    for contig_record in contig_parse:
        contig_id = contig_record.id
        if contig_id in virus2tax.keys():
            virus_contig.append(contig_record)
    SeqIO.write(virus_contig, output_contig, 'fasta')

def generate_virus_gene(orf_file, virus2tax, output_gene):
    orf_parse = SeqIO.parse(orf_file, 'fasta')
    virus_gene = []
    for orf_record in orf_parse:
        orf_id = orf_record.id
        contig_id = orf_id.rsplit("_", 1)[0]
        if contig_id in virus2tax.keys():
            virus_gene.append(orf_record)
    SeqIO.write(virus_gene, output_gene, 'fasta')

def get_clade(taxid, clade_list):
    clade_name = []
    clade_taxid = []
    lineages = ncbi.get_lineage(taxid)
    rank_list = ncbi.get_rank(lineages)
    for lineage in lineages:
        if rank_list[lineage] in clade_list:
            name = ncbi.get_taxid_translator([lineage])[lineage]
            clade_name.append(name)
            clade_taxid.append(str(lineage))
    return '|'.join(clade_name), '|'.join(clade_taxid)

def generate_CAMI(virus_abundance, cami_output):
    outf = open(cami_output, 'w')
    outf.write('@SampleID:simulation\n@Version:0.10.0\n@Ranks:superkingdom|phylum|class|order|family|genus|species|strain\n@@TAXID\tRANK\tTAXPATH\tTAXPATHSN\tPERCENTAGE\n')
    clade_list = ['superkingdom', 'phylum', 'class', 'order', 'family', 'genus', 'species', 'strain']
    clade2num = {'superkingdom':0, 'phylum':1, 'class':2, 'order':3, 'family':4, 'genus':5, 'species':6, 'strain':7}
    abundance_table = pd.DataFrame(columns=['RANK', 'TAXPATH', 'TAXPATHSN','PERCENTAGE', 'number'])
    for taxid in virus_abundance.keys():
        percentage = virus_abundance[taxid]['percentage']
        lineages = ncbi.get_lineage(taxid)
        rank_list = ncbi.get_rank(lineages)
        for lineage in lineages:
            if rank_list[lineage] in clade_list:
                if lineage not in abundance_table.index:
                    clade_name, clade_taxid = get_clade(lineage, clade_list)
                    abundance_table.loc[lineage] = ({
                        'RANK': rank_list[lineage],
                        'TAXPATH': clade_taxid,
                        'TAXPATHSN': clade_name,
                        'PERCENTAGE': percentage,
                        'number': clade2num[rank_list[lineage]]
                    })
                else:
                    abundance_table.loc[lineage, 'PERCENTAGE'] += percentage
    
    #print(abundance_table)
    abundance_table.sort_values(by=['number', 'PERCENTAGE'], ascending=[True, False], inplace=True)
    #print(abundance_table)
    for taxid in abundance_table.index:
        rank = abundance_table.loc[taxid, 'RANK']
        taxpath = abundance_table.loc[taxid, 'TAXPATH']
        taxpathsn = abundance_table.loc[taxid, 'TAXPATHSN']
        percentage = abundance_table.loc[taxid, 'PERCENTAGE']
        outf.write('\t'.join([str(taxid), rank, taxpath, taxpathsn, str(percentage)])+ '\n')
    outf.close()
    

def generate_report(virus_anno, nonvirus_anno, ORF_faa, contig, quant, cluster_rep, cluter_file, one_minus_r, report_dir, report_log, processes):
    try:
        if quant != None:
            ORF_abundance = pd.read_csv(quant, sep='\t')
        else:
            ORF_abundance = pd.DataFrame()
            '''virus2tax, virus_abundance, organism_abundance = classify_multi.classify(
                args.virus_anno, args.nonvirus_anno,
                ORF_abundance,
                args.orf,
                args.one_minus_r,
                int(args.processes)
            )'''
        virus2tax, ambiguous2tax, virus2multi_tax, virus_abundance, organism_abundance = classify_radical.classify(
            virus_anno, nonvirus_anno,
            ORF_abundance,
            ORF_faa,
            cluster_rep,
            cluter_file,
            one_minus_r,
            int(processes)
        )
        if virus_abundance:
            dna_virus, rna_virus, unknown_virus = virus_divide.get_virus_type(
                virus2tax,
                virus_abundance
            )
            generate_CAMI(
                virus_abundance,
                os.path.join(report_dir,'virus.cami.profile')
            )
            write_output(
                dna_virus['tax'],
                dna_virus['abundance'],
                contig, ORF_faa,
                os.path.join(report_dir, 'dna_virus'),
                'dna_virus',
                report_log
            )
            write_output(
                rna_virus['tax'],
                rna_virus['abundance'],
                contig, ORF_faa,
                os.path.join(report_dir, 'rna_virus'),
                'rna_virus',
                report_log
            )
            write_output(
                unknown_virus['tax'],
                unknown_virus['abundance'],
                contig, ORF_faa,
                os.path.join(report_dir, 'unknown_virus'),
                'unknown_virus',
                report_log
            )
        write_output(
            virus2tax,
            virus_abundance,
            contig, ORF_faa,
            os.path.join(report_dir, 'all_virus'),
            'virus',
            report_log
        )
        write_output(
            ambiguous2tax,
            None,
            contig, ORF_faa,
            os.path.join(report_dir, 'ambiguous'),
            'ambiguous',
            report_log
        )
        write_output(
            virus2multi_tax,
            None,
            contig, ORF_faa,
            os.path.join(report_dir, 'virus_multi'),
            'virus_multi',
            report_log
        )

        return True
    except Exception as e:
        print("ERROR: Generate Report Fail!")
        logger.error(traceback.format_exc())
        logger.error("ERROR: Generate Report Fail!")
        sys.exit()




    