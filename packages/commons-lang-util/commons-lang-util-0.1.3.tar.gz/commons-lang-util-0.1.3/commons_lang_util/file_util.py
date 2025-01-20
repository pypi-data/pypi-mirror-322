import csv
import json
import os
import re
import shutil
import sys
import zipfile
from typing import List, Dict
from datetime import datetime

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.workbook import Workbook


def create_empty_file(file):
    append_as_text_line(file, "no content!")


def list_all_files_in_dir(dir_path,
                          postfixes: List[str] = None,
                          prefixes: List[str] = None,
                          exclude_postfixes: List[str] = None,
                          exclude_prefixes: List[str] = None,
                          ) -> List[str]:
    def match_postfixes(file, postfixes):
        for postfix in postfixes:
            if file.endswith(postfix): return True
        return False

    def match_prefixes(file, prefixes):
        for prefix in prefixes:
            if file.startswith(prefix): return True
        return False

    all_files = []
    for root, dirs, files in os.walk(dir_path):
        for file in files:
            if prefixes is not None and not match_prefixes(file, prefixes):
                continue
            if postfixes is not None and not match_postfixes(file, postfixes):
                continue
            if exclude_postfixes is not None and match_postfixes(file, exclude_postfixes):
                continue
            if exclude_prefixes is not None and match_prefixes(file, exclude_prefixes):
                continue
            file_path = os.path.join(root, file)
            all_files.append(file_path)
    return all_files


def copy_dir(source_folder, destination_folder):
    try:
        # 拷贝文件夹及文件夹内的所有文件
        shutil.copytree(source_folder, destination_folder)
    except Exception as e:
        raise e


def copy_file(source_file, destination_file):
    try:
        # 拷贝文件
        shutil.copy(source_file, destination_file)
    except Exception as e:
        raise e


def remove_dir(dir_path):
    if os.path.exists(dir_path):
        shutil.rmtree(dir_path)


def remove_file(file):
    if os.path.exists(file):
        os.remove(file)


def read_lines(file):
    with open(file, "r", encoding="utf8") as f:
        return [line.strip() for line in f.readlines() if line.strip()]


def read_string(file):
    with open(file, "r", encoding="utf8") as f:
        return f.read()


def read_json(file):
    with open(file, "r", encoding="utf8") as f:
        return json.loads(f.read())


def read_xlsx_deprecated(file, ignore_first_line: bool, output_headers: List[str]) -> List[Dict[str, any]]:
    lines = []
    min_row = 2 if ignore_first_line else 1
    workbook = openpyxl.load_workbook(file)
    # 选择要读取的工作表
    sheet = workbook.active  # 或者使用 workbook['工作表名称']

    # 遍历行
    for row in sheet.iter_rows(min_row=min_row, values_only=True):  # 从第二行开始遍历
        line = {}
        if row[0] is None: break
        for i in range(len(output_headers)):
            line[output_headers[i]] = row[i]
        lines.append(line)

    # 关闭XLSX文件
    workbook.close()
    return lines


def read_xlsx(file):
    workbook = openpyxl.load_workbook(file)
    # 选择要读取的工作表
    sheet = workbook.active  # 或者使用 workbook['工作表名称']

    # 获取表头
    headers = [cell.value for cell in sheet[1]]
    lines = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # 从第二行开始遍历
        line = {}
        if row[0] is None: break
        for i in range(len(headers)):
            value = row[i]
            str_value = None if value is None or str(value) == "NULL" else str(value)
            line[headers[i]] = str_value
        lines.append(line)

        # 关闭XLSX文件
    workbook.close()
    return lines


def read_csv(file, encoding="utf-8", delimiter=',') -> List[Dict[str, any]]:
    with open(file, "r", encoding=encoding) as file:
        csv_reader = csv.reader(file, delimiter=delimiter)
        headers = next(csv_reader)
        # print(_headers)
        lines = []
        for row in csv_reader:
            line = {}
            for i in range(len(headers)):
                line[headers[i]] = row[i]
            lines.append(line)
        return lines


def read_csv_deprecated(file, output_headers: List[str]) -> List[Dict[str, any]]:
    with open(file, "r", encoding="utf-8") as file:
        csv_reader = csv.reader(file)
        _header = next(csv_reader)
        lines = []
        for row in csv_reader:
            line = {}
            for i in range(len(output_headers)):
                line[output_headers[i]] = row[i]
            lines.append(line)
        return lines


def make_dir(file):
    dir_path = os.path.dirname(file)
    if dir_path != "" and not os.path.exists(dir_path):
        os.makedirs(dir_path)


def write_lines(file, lines):
    make_dir(file)
    with open(file, "w", encoding="utf-8") as f:
        for line in lines:
            f.write(line + "\n")


def write_map(file, map):
    make_dir(file)
    with open(file, "w") as f:
        for k, v in map.items():
            f.write("%s %s \n" % (str(k), str(v)))


def write_str(file, str):
    make_dir(file)
    with open(file, "w", encoding="utf-8") as f:
        f.write(str)


def write_as_json(file, obj, encoding="utf-8", pretty=True):
    make_dir(file)
    with open(file, "w", encoding=encoding) as f:
        f.write(json.dumps(obj, ensure_ascii=False, indent=4 if pretty else None))


def write_as_xlsx_with_headers(file, list_of_maps, headers, remove_illegal_chars=False):
    make_dir(file)
    lines = []
    for map in list_of_maps:
        line = []
        for header in headers:
            value = map.get(header, "-")
            if type(value) == list:
                if len(value) == 0:
                    value = "-"
                else:
                    value = " | ".join([str(e) for e in value])
            if remove_illegal_chars and type(value) == str:
                value = re.sub(ILLEGAL_CHARACTERS_RE, "", value)
            line.append(value)
        lines.append(line)
    # with open(file, "w", encoding=encoding, newline="") as f:
    #     csv_writer = csv.writer(f)
    #     for line in lines:
    #         csv_writer.writerow(line)
    workbook = Workbook()
    sheet = workbook.active

    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, row_data in enumerate(lines, 2):
        for col_num, raw_value in enumerate(row_data, 1):
            cell_value = raw_value
            if raw_value is not None and str(raw_value).startswith("="): cell_value = "'%s" % raw_value
            sheet.cell(row=row_num, column=col_num, value=cell_value)
            # print("write: row: %s, column: %s, value: %s" % (row_num, col_num, value))

    workbook.save(file)


def __get_hour_minute_str():
        now = datetime.now()
        formatted_time = f"{now.hour:02}-{now.minute:02}"
        return formatted_time


def write_as_xlsx(file, list_of_maps, remove_illegal_chars=False, time_postfix=False):
    """
    :param file:
    :param list_of_maps:
    :param remove_illegal_chars: 有些特殊字符会导致写入xlsx报错.
    :return:
    """
    
    

    if time_postfix:
        splits = file.rsplit(".", 1)
        file_name_with_path = splits[0]
        postfix = splits[1]
        file = "%s_%s.%s" % (file_name_with_path, __get_hour_minute_str(), postfix)

    if list_of_maps is None or len(list_of_maps) == 0:
        create_empty_file(file)
        return

    def get_headers(_list_of_maps):
        _headers = []
        first_line = _list_of_maps[0]
        for header in first_line.keys():
            _headers.append(header)
        return _headers

    # 根据内容自动生成csv headers, 依赖dict遍历是有序的, 因此要求python版本>=3.7
    if sys.version_info < (3, 7):
        raise Exception("Require python version >=3.7")

    headers = get_headers(list_of_maps)
    write_as_xlsx_with_headers(file, list_of_maps, headers, remove_illegal_chars)


def write_as_csv_with_headers(file, list_of_maps, headers, encoding="utf-8", delimiter=','):
    make_dir(file)
    lines = []
    lines.append(headers)
    for map in list_of_maps:
        line = []
        for header in headers:
            value = map.get(header, "-")
            if type(value) == list or type(value) == set:
                if len(value) == 0:
                    value = "-"
                else:
                    value = " | ".join([str(e) for e in value])
            line.append(value)
        lines.append(line)
    with open(file, "w", encoding=encoding, newline="") as f:
        csv_writer = csv.writer(f, delimiter=delimiter)
        for line in lines:
            csv_writer.writerow(line)


def write_as_csv(file, list_of_maps, encoding="utf-8", delimiter=','):
    if list_of_maps is None or len(list_of_maps) == 0:
        create_empty_file(file)
        return

    def get_headers(_list_of_maps):
        _headers = []
        first_line = _list_of_maps[0]
        for header in first_line.keys():
            _headers.append(header)
        return _headers

    # 根据内容自动生成csv headers, 依赖dict遍历是有序的, 因此要求python版本>=3.7
    if sys.version_info < (3, 7):
        raise Exception("Require python version >=3.7")

    headers = get_headers(list_of_maps)
    write_as_csv_with_headers(file, list_of_maps, headers, encoding=encoding, delimiter=delimiter)


def write_to_zip(folder_path, zip_file_path):
    make_dir(zip_file_path)
    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, _, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                zipf.write(file_path, os.path.relpath(file_path, folder_path))


def append_as_text_line(file, text: str):
    make_dir(file)
    with open(file, "a", encoding="utf-8") as f:
        f.write(text + "\n")


def append_as_json_string_line(file, obj):
    make_dir(file)
    with open(file, "a", encoding="utf-8") as f:
        f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def extrac_desc_from_file_name_parentheses(file):
    """
    从文件名中提取小括号之内的文本
    """
    file_name = os.path.basename(file)
    left_index = file_name.index("(")
    right_index = file_name.index(")")
    return file_name[left_index + 1: right_index]


def exists(path):
    return os.path.exists(path)


def escape_file_name(file_name):
    # 文件名中不能出现以下特殊字符
    special_chars = ["\\", "/", ":", "*", "?", "\"", "<", ">", "|"]
    for char in special_chars:
        file_name = file_name.replace(char, "_")
    return file_name


def split_xlsx_by_line(input_file, output_file_prefix, split_line_num=1000):
    lines = read_xlsx(input_file)
    split_count = int(len(lines) / split_line_num) + 1
    for i in range(split_count):
        per_file_lines = lines[i * split_line_num:(i + 1) * split_line_num]
        output_file = "%s_split_%s.xlsx" % (output_file_prefix, i)
        write_as_xlsx(output_file, per_file_lines)

    print("split complete.")


def merge_xlsx(input_dir, output_file):
    def get_creation_time(file_path):
        return os.path.getctime(file_path)

    # 排除"~$"开头的文件, 这个是打开xlsx文件时临时生成的.
    files = list_all_files_in_dir(input_dir, postfixes=["xlsx"], exclude_prefixes=["~$"])
    files = sorted(files, key=lambda x: get_creation_time(x))
    all_lines = []
    for file in files:
        print("merge file: %s" % file)
        all_lines.extend(read_xlsx(file))
    print("after merge size: %s" % len(all_lines))
    write_as_xlsx(output_file, all_lines)

