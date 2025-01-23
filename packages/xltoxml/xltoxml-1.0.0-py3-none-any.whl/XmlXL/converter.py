# my_utility/converter.py

import xml.etree.ElementTree as ET
from openpyxl import Workbook

class XmlToExcel:
    def __init__(self, xml_file, excel_file):
        self.xml_file = xml_file
        self.excel_file = excel_file

    def parse_element(self, element, ws, row, col):
        """Recursively parse XML and write hierarchy to Excel."""
        # Write the element tag in the appropriate column
        ws.cell(row=row, column=col, value=element.tag)
        
        # Recursively handle child elements
        next_row = row + 1
        for child in element:
            # Write child element tags in the next column
            next_row = self.parse_element(child, ws, next_row, col + 1)  # Move child elements one column to the right
        
        # Write the data (if any) after the last child element of the current element
        if element.text and element.text.strip():
            # Now, we should write the text in the column after the last child of the current element's tag
            ws.cell(row=row, column=col + 1, value=element.text.strip())
        
        return next_row

    def convert(self):
        """Converts the provided XML to an Excel file."""
        tree = ET.parse(self.xml_file)
        root = tree.getroot()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "XML_Data"
        
        # Start writing data from row 2
        self.parse_element(root, ws, row=1, col=1)
        wb.save(self.excel_file)
        print(f"XML converted to Excel and saved as {self.excel_file}.")

