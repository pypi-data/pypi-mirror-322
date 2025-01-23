import json
from openpyxl import Workbook

class JsonToExcel:
    def __init__(self, json_file, excel_file):
        self.json_file = json_file
        self.excel_file = excel_file
        self.data = None
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.current_row = 1

    def read_json(self):
        """Reads the JSON file and loads data."""
        with open(self.json_file, 'r') as file:
            self.data = json.load(file)

    def write_hierarchy(self, data, col=1):
        """Writes the hierarchical data to Excel."""
        if isinstance(data, dict):
            for key, value in data.items():
                self.sheet.cell(row=self.current_row, column=col, value=key)
                self.current_row += 1
                self.write_hierarchy(value, col + 1)
        
        elif isinstance(data, list):
            for item in data:
                self.write_hierarchy(item, col)
        
        else:
            self.sheet.cell(row=self.current_row, column=col, value=data)
            self.current_row += 1

    def save_to_excel(self):
        """Saves the data to an Excel file."""
        self.workbook.save(self.excel_file)

    def convert(self):
        """Full conversion process."""
        if self.data:
            self.write_hierarchy(self.data)
            self.save_to_excel()

if __name__ == "__main__":
    # File paths
    json_file = "/home/maheshreddy/Desktop/jsontoexcelheirachystrucure/example_2.json"
    excel_file = "/home/maheshreddy/Desktop/jsontoexcelheirachystrucure/out.xlsx"

    # Create the converter and process the JSON
    converter = JsonToExcel(json_file, excel_file)
    converter.read_json()  # Read the JSON file
    converter.convert()  # Perform the conversion if the data is loaded

