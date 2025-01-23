import re
from openpyxl import load_workbook
import xml.etree.ElementTree as ET

class ExcelToXML:
    def __init__(self, excel_file, xml_file):
        self.excel_file = excel_file
        self.xml_file = xml_file

    def sanitize_tag(self, tag):
        """Sanitize the tag to ensure it's a valid XML element name."""
        # Remove invalid characters (anything that is not a letter, digit, underscore, or hyphen)
        tag = re.sub(r'[^A-Za-z0-9_-]', '_', tag)
        
        # Ensure the tag doesn't start with a digit (invalid in XML)
        if tag[0].isdigit():
            tag = f'_{tag}'
        
        return tag

    def build_xml(self, ws, row, col):
        """Recursively build XML from Excel."""
        tag = ws.cell(row=row, column=col).value
        if not tag:
            print(f"Skipping empty cell at row={row}, col={col}")
            return None, row
        
        # Sanitize the tag to ensure it is a valid XML name
        tag = self.sanitize_tag(tag)
        element = ET.Element(tag)
        print(f"Processing tag: {tag} at row={row}, col={col}")
        
        next_row = row + 1
        while next_row <= ws.max_row:
            child_tag = ws.cell(row=next_row, column=col + 1).value
            if not child_tag:
                break  # No child elements found, stop recursion
            
            # Recursively build the child element
            child_element, next_row = self.build_xml(ws, next_row, col + 1)
            if child_element is not None:
                element.append(child_element)
        
        return element, next_row

    def convert(self):
        """Convert the Excel file to XML."""
        wb = load_workbook(self.excel_file)
        ws = wb.active

        root, _ = self.build_xml(ws, row=1, col=1)

        if root is None:
            print("Error: No valid data found in Excel to convert.")
            return
        
        tree = ET.ElementTree(root)
        tree.write(self.xml_file, encoding="utf-8", xml_declaration=True)
        print(f"Excel converted back to XML and saved as {self.xml_file}.")

# Usage
excel_to_xml = ExcelToXML("output.xlsx", "last.xml")
excel_to_xml.convert()




